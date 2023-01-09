import openpyxl


#Same Data
#file="C:\\data\\testdata.xlsx"
#workbook=openpyxl.load_workbook(file)
#sheet=workbook.active # to get from active sheet instead sheet=workbook.["Sheet1"]
#for r in range(1,6):
#    for c in range(1,4):
#       sheet.cell(r,c).value="Welcome"
#workbook.save(file)


#multiple Data

file="C:\\data\\testdata1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active # to get from active sheet instead sheet=workbook.["Sheet1"]
sheet.cell(1,1).value=123
sheet.cell(1,2).value="shibham"
sheet.cell(1,3).value="Engineer"

sheet.cell(2,1).value=567
sheet.cell(2,2).value="Linge"
sheet.cell(2,3).value="QA"


workbook.save(file)


