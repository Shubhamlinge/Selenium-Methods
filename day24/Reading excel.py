import openpyxl


#File -> workbook -> sheet -> rows → column

file="C:\data\data.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row #count number of rows inexcel sheet
cols=sheet.max_column #count number of column inexcel sheet

# Reading all the rows and column from excel sheet

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='         ')
    print()



